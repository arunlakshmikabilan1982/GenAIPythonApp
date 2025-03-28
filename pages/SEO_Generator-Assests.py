from langchain_community.document_loaders import PyPDFLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_pinecone import PineconeVectorStore
from langchain_community.embeddings import HuggingFaceEmbeddings
from langchain_google_genai import ChatGoogleGenerativeAI
import google.generativeai as genai
from pinecone import Pinecone, ServerlessSpec
from langchain.chains import ConversationalRetrievalChain
from langchain_core.prompts import ChatPromptTemplate, SystemMessagePromptTemplate, HumanMessagePromptTemplate, AIMessagePromptTemplate
from langchain_core.prompts import PromptTemplate
# from environment import PINECONE_SEO_INDEX, GEMINI_API_KEY, PINECONE_API_KEY
from PIL import Image
import langchain
langchain.verbose = False
import asyncio
# Create a new event loop
loop = asyncio.new_event_loop()
# Set the event loop as the current event loop
asyncio.set_event_loop(loop)
import streamlit as st
import openpyxl
import tempfile
import requests
import io, json
from Navigation import sidebar

sidebar()

GEMINI_API_KEY = st.secrets["GEMINI_API_KEY"]
PINECONE_API_KEY = st.secrets["PINECONE_API_KEY"]
PINECONE_SEO_INDEX = st.secrets["PINECONE_SEO_INDEX"]
 
# pineconeindex = pc.Index(PINECONE_SEO_INDEX)
pc = Pinecone(api_key=PINECONE_API_KEY)
genai.configure(api_key=GEMINI_API_KEY)
def imagetotext(imageurl):
    #   converted_string = base64.b64encode(imageurl.read()) 
    #   print(converted_string)
      imageurlwords = imageurl.replace('/', " ").replace("."," ").replace("-"," ")
      print(imageurlwords)

      # Headers to mimic a browser visit
      headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Accept-Language": "en-US,en;q=0.9",
            "Accept-Encoding": "gzip, deflate, br",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
            "Referer": "https://www.google.com/",
            "Connection": "keep-alive"
        } 
      
      try:
            result = requests.get(imageurl, headers=headers, timeout=10)
      except requests.exceptions.ConnectionError:
            print("Site not rechable", imageurl)

      image = Image.open(io.BytesIO(result.content))
    #   image = Image.open(requests.get(imageurl, stream=True).raw)
      prompt_template: str = r"""
          Given a specific context, Generate SEO Metadata details in 'Json Format' as per provided template, use valid words from provided {imageurlwords} while generating the SEO metadata details, and Strictly Use below Template:
          FileName: filename for the image, which will define the image purpose
          Title: eyecatching title for the image, should be atleast 5 words
          Description: All the important details like configuration,features about the image, Not to exceed 50 words, should be in paragraph
          Keywords: keywords for the image provided, should be string value

          role: Metadata Content Creator
          question: {question}
          imageurlwords: {imageurlwords}
          """     
      prompt = PromptTemplate.from_template(template=prompt_template)
      input = "follow template and role provided"
      prompt_formatted_str: str = prompt.format(question=input, imageurlwords = {imageurlwords})
      model = genai.GenerativeModel('gemini-1.5-pro-latest')
      response = model.generate_content([prompt_formatted_str,image])  
      return response.text

def load_pdf_documents(pdf_url):
    loader = PyPDFLoader(pdf_url, extract_images=True)
    documents = loader.load()
    return documents

model = ChatGoogleGenerativeAI(model="gemini-1.0-pro-latest",
                            google_api_key=GEMINI_API_KEY,
                            temperature=0.2,
                            convert_system_message_to_human=True)

def split_documents(documents):
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=900, chunk_overlap=50,separators=[" ", ",", "\n"])
    texts = text_splitter.split_documents(documents)
    return texts

def embeddings_on_pinecone(texts):
    embeddings = HuggingFaceEmbeddings()
    if(PINECONE_SEO_INDEX in pc.list_indexes().names()):
        print(pc.describe_index(PINECONE_SEO_INDEX))
        pc.delete_index(PINECONE_SEO_INDEX)
    pc.create_index(
    name=PINECONE_SEO_INDEX,
    dimension=768,
    metric="cosine",
    spec=ServerlessSpec(
        cloud="aws",
        region="us-east-1"
    )
    )
    PineconeVectorStore.from_documents(texts, embeddings, index_name=PINECONE_SEO_INDEX)

def retriever_existingdb():
    embeddings = HuggingFaceEmbeddings()
    vectorstore = PineconeVectorStore.from_existing_index(index_name=PINECONE_SEO_INDEX, embedding=embeddings)
    retriever = vectorstore.as_retriever(
    search_type="similarity_score_threshold", search_kwargs={"score_threshold": 0.5}
    )
    return retriever

def query_llm(retriever, query):
    general_system_template = r""" 
    Given a specific context, Generate SEO Metadata details in 'Json Format' as per provided template, Strictly Use below Template:
    FileName: filename for the PDF document, which will define the pdf document purpose
    Title: eyecatching title for the image, should be atleast 5 words
    Description: All the important details like configuration,features about the image, Not to exceed 50 words, should be in paragraph
    Keywords: keywords for the image provided, should be string value
    ----
    {context}
    ----
    """

    general_ai_template = "role:SEO metadata content creator"
    general_user_template = "Question:```{query}```"
    messages = [
            SystemMessagePromptTemplate.from_template(general_system_template),
            HumanMessagePromptTemplate.from_template(general_user_template),
            AIMessagePromptTemplate.from_template(general_ai_template)
               ]
    qa_prompt = ChatPromptTemplate.from_messages( messages )
    # model = ChatGoogleGenerativeAI(model="gemini-1.0-pro-latest",
    #                         google_api_key=GEMINI_API_KEY,
    #                         temperature=0.2,
    #                         convert_system_message_to_human=True)
    qa = ConversationalRetrievalChain.from_llm(
            model,
            retriever=retriever,
            chain_type="stuff",
            combine_docs_chain_kwargs={'prompt': qa_prompt}
        )
    result = qa({"question": query, "query": query, "chat_history": ""})
    result = result["answer"]
    return result

def process_pdf_documents(pdffile):
    print(pdffile)
    documents = load_pdf_documents(pdffile)
    texts = split_documents(documents)
    embeddings_on_pinecone(texts)
    retriever  =  retriever_existingdb()
    query = "generate SEO Metadata details as per requested template"
    results = query_llm(retriever, query)
    return results

def Image_from_url(image_url):
    results = imagetotext(image_url)
    print(results)
    return results
 
def PdfText_from_url(pdf_url):
    results = process_pdf_documents(pdf_url)
    print(results)
    return results

# Function to update Excel file with keywords
def update_excel_with_seo(excel_file, filename, title, description, keywords):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    filenamecount = 2  # Start from the second row (assuming the first row is headers)
    for row in ws.iter_rows(min_row=2, max_row=len(filename)+1, min_col=3, max_col=3):  # Iterate over column "C"
        row[0].value = filename[filenamecount - 2]  # Update cell value with filename
        filenamecount += 1

    titlecount = 2  # Start from the second row (assuming the first row is headers)
    for row in ws.iter_rows(min_row=2, max_row=len(title)+1, min_col=4, max_col=4):  # Iterate over column "C"
        row[0].value = title[titlecount - 2]  # Update cell value with title
        titlecount += 1
   
    descriptioncount = 2
    for row in ws.iter_rows(min_row=2, max_row=len(description)+1, min_col=5, max_col=5):  # Iterate over column "D"
        row[0].value = description[descriptioncount - 2]  # Update cell value with description
        descriptioncount += 1

    keywordscount = 2
    for row in ws.iter_rows(min_row=2, max_row=len(keywords)+1, min_col=6, max_col=6):  # Iterate over column "E"
        row[0].value = keywords[keywordscount - 2]  # Update cell value with keywords
        keywordscount += 1

    # Save the workbook to a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
        wb.save(tmp_file.name)
        return tmp_file.name
 
def uploadfile(uploaded_file): 
 if uploaded_file is not None:
    if uploaded_file.name.endswith('.xlsx'):  # Check if the uploaded file is an Excel file
        title = []
        filename = []
        keywords = []
        description = []
        wb = openpyxl.load_workbook(uploaded_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=2):  # Iterate over rows
            file_type, url = row[0].value, row[1].value
            if file_type == "Pdf":
                results = process_pdf_documents(url)
                print(results)
                results = results.replace("```json","").replace("```","")
                jsonvalue = json.loads(results)
                filename.append(jsonvalue["FileName"])
                title.append(jsonvalue["Title"])
                description.append(jsonvalue["Description"])
                keywords.append(jsonvalue["Keywords"])

            elif file_type == "Image":
                print(url)
                results = imagetotext(url)
                print(results)
                results = results.replace("```json","").replace("```","")
                jsonvalue = json.loads(results)
                filename.append(jsonvalue["FileName"])
                title.append(jsonvalue["Title"])
                description.append(jsonvalue["Description"])
                keywords.append(jsonvalue["Keywords"])


        if filename and title and description and keywords:
            updated_file_path = update_excel_with_seo(uploaded_file, filename, title, description, keywords)
            st.success("The SEO metadata has been successfully updated.")
 
            # Provide a download link for the updated Excel file
            with open(updated_file_path, 'rb') as f:
                data = f.read()
                st.download_button(
                    label="Download Updated Excel File",
                    data=data,
                    file_name="SEO_Assets.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.error("Please ensure that the file you upload is a valid Excel file.")

def main():
    st.title("SEO Metadata Generator - Assets")
    uploaded_file = st.file_uploader("Please select the file you would like to upload",type=["xlsx"])
    if uploaded_file is not None:
        with st.spinner("Generating SEO MetaData..."):
            uploadfile(uploaded_file)
        
if __name__ == "__main__":
    main()
